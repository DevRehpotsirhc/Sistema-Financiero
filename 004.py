import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import sqlite3
import hashlib
import datetime
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl

DB_FILE = 'finanzas.db'

# ---------------------------------------
# --- BASE DE DATOS ---------------------
# ---------------------------------------

def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.execute('PRAGMA foreign_keys = ON')
    return conn

def init_db():
    conn = get_db_connection()
    c = conn.cursor()

    # Tabla usuarios
    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            apellido TEXT NOT NULL,
            cedula TEXT NOT NULL UNIQUE,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('master', 'estandar'))
        )
    ''')

    # Tabla transacciones
    c.execute('''
        CREATE TABLE IF NOT EXISTS transacciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL,
            usuario_id INTEGER NOT NULL,
            tipo_mov TEXT NOT NULL CHECK(tipo_mov IN ('entrada', 'salida')),
            categoria TEXT NOT NULL,
            descripcion TEXT,
            moneda TEXT NOT NULL CHECK(moneda IN ('bs', 'usd')),
            monto REAL NOT NULL,
            FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
        )
    ''')

    conn.commit()
    conn.close()

def hash_password(password):
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def check_first_user_exists():
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM usuarios')
    exists = c.fetchone()[0] > 0
    conn.close()
    return exists

def crear_usuario(nombre, apellido, cedula, username, password, tipo):
    conn = get_db_connection()
    c = conn.cursor()
    try:
        c.execute('INSERT INTO usuarios (nombre, apellido, cedula, username, password, tipo) VALUES (?, ?, ?, ?, ?, ?)',
                  (nombre, apellido, cedula, username, hash_password(password), tipo))
        conn.commit()
        return True, None
    except sqlite3.IntegrityError as e:
        return False, str(e)
    finally:
        conn.close()

def validar_login(username, password):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT id, nombre, apellido, tipo, password FROM usuarios WHERE username = ?', (username,))
    row = c.fetchone()
    conn.close()
    if row and hash_password(password) == row[4]:
        return {
            'id': row[0],
            'nombre': row[1],
            'apellido': row[2],
            'tipo': row[3],
            'username': username
        }
    else:
        return None

# ---------------------------------------
# --- FUNCIONES DE TRANSACCIONES --------
# ---------------------------------------

def registrar_transaccion(usuario_id, tipo_mov, categoria, descripcion, moneda, monto):
    now = datetime.datetime.now()
    fecha = now.strftime('%Y-%m-%d')
    hora = now.strftime('%H:%M:%S')
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''
        INSERT INTO transacciones (fecha, hora, usuario_id, tipo_mov, categoria, descripcion, moneda, monto)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (fecha, hora, usuario_id, tipo_mov, categoria, descripcion, moneda, monto))
    conn.commit()
    conn.close()

def obtener_transacciones(filtros=None):
    # filtros es dict con claves: fecha_inicio, fecha_fin, usuario_id, tipo_mov, categoria, moneda, texto_busqueda
    conn = get_db_connection()
    c = conn.cursor()
    query = 'SELECT t.id, t.fecha, t.hora, u.username, t.tipo_mov, t.categoria, t.descripcion, t.moneda, t.monto FROM transacciones t JOIN usuarios u ON t.usuario_id = u.id WHERE 1=1'
    params = []

    if filtros:
        if 'fecha_inicio' in filtros and filtros['fecha_inicio']:
            query += ' AND fecha >= ?'
            params.append(filtros['fecha_inicio'])
        if 'fecha_fin' in filtros and filtros['fecha_fin']:
            query += ' AND fecha <= ?'
            params.append(filtros['fecha_fin'])
        if 'usuario_id' in filtros and filtros['usuario_id']:
            query += ' AND usuario_id = ?'
            params.append(filtros['usuario_id'])
        if 'tipo_mov' in filtros and filtros['tipo_mov'] in ('entrada', 'salida'):
            query += ' AND tipo_mov = ?'
            params.append(filtros['tipo_mov'])
        if 'categoria' in filtros and filtros['categoria']:
            query += ' AND categoria LIKE ?'
            params.append('%' + filtros['categoria'] + '%')
        if 'moneda' in filtros and filtros['moneda'] in ('bs', 'usd'):
            query += ' AND moneda = ?'
            params.append(filtros['moneda'])
        if 'texto' in filtros and filtros['texto']:
            query += ' AND (descripcion LIKE ? OR categoria LIKE ?)'
            params.extend(['%' + filtros['texto'] + '%', '%' + filtros['texto'] + '%'])

    query += ' ORDER BY fecha DESC, hora DESC'
    c.execute(query, params)
    resultados = c.fetchall()
    conn.close()
    return resultados

def calcular_balance(usuario_id=None):
    conn = get_db_connection()
    c = conn.cursor()
    if usuario_id:
        c.execute('SELECT moneda, SUM(CASE WHEN tipo_mov="entrada" THEN monto ELSE -monto END) FROM transacciones WHERE usuario_id=? GROUP BY moneda', (usuario_id,))
    else:
        c.execute('SELECT moneda, SUM(CASE WHEN tipo_mov="entrada" THEN monto ELSE -monto END) FROM transacciones GROUP BY moneda')
    rows = c.fetchall()
    conn.close()
    balances = {row[0]: row[1] if row[1] is not None else 0 for row in rows}
    # asegurar que estén ambos monedas
    if 'bs' not in balances: balances['bs'] = 0
    if 'usd' not in balances: balances['usd'] = 0
    return balances

# ---------------------------------------
# --- REPORTES PDF ----------------------
# ---------------------------------------

def generar_reporte_pdf(ruta_pdf, transacciones, usuario_actual):
    c = canvas.Canvas(ruta_pdf, pagesize=letter)
    width, height = letter
    margen = 50
    y = height - margen

    c.setFont("Helvetica-Bold", 16)
    c.drawString(margen, y, "Reporte de Transacciones")
    y -= 25
    c.setFont("Helvetica", 10)
    c.drawString(margen, y, f"Generado por: {usuario_actual['nombre']} {usuario_actual['apellido']} ({usuario_actual['username']})")
    y -= 15
    c.drawString(margen, y, f"Fecha generación: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 25

    c.setFont("Helvetica-Bold", 9)
    encabezados = ["ID", "Fecha", "Hora", "Usuario", "Tipo", "Categoría", "Descripción", "Moneda", "Monto"]
    posiciones = [margen, margen+30, margen+85, margen+140, margen+210, margen+270, margen+350, margen+450, margen+490]
    for i, encabezado in enumerate(encabezados):
        c.drawString(posiciones[i], y, encabezado)
    y -= 15
    c.setFont("Helvetica", 8)

    for t in transacciones:
        if y < 50:
            c.showPage()
            y = height - margen
            c.setFont("Helvetica-Bold", 9)
            for i, encabezado in enumerate(encabezados):
                c.drawString(posiciones[i], y, encabezado)
            y -= 15
            c.setFont("Helvetica", 8)
        # Datos
        c.drawString(posiciones[0], y, str(t[0]))
        c.drawString(posiciones[1], y, t[1])
        c.drawString(posiciones[2], y, t[2])
        c.drawString(posiciones[3], y, t[3])
        c.drawString(posiciones[4], y, t[4])
        c.drawString(posiciones[5], y, t[5])
        desc = t[6] if t[6] else ""
        c.drawString(posiciones[6], y, desc[:18])
        c.drawString(posiciones[7], y, t[7])
        c.drawRightString(posiciones[8]+40, y, f"{t[8]:,.2f}")
        y -= 12

    c.save()

# ---------------------------------------
# --- IMPORTAR/EXPORTAR EXCEL -----------
# ---------------------------------------

def exportar_a_excel(ruta_excel, transacciones):
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws:
        ws.title = "Transacciones"
        encabezados = ["ID", "Fecha", "Hora", "Usuario", "Tipo", "Categoría", "Descripción", "Moneda", "Monto"]
        ws.append(encabezados)
        for t in transacciones:
            ws.append(t)
        wb.save(ruta_excel)

def importar_desde_excel(ruta_excel):
    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb.active
    if ws:
        filas = list(ws.iter_rows(min_row=2, values_only=True))
        # Se espera que las columnas sean las mismas que exportar
        # Ignoramos ID porque es autoincremental
        transacciones = []
        for fila in filas:
            try:
                fecha, hora, usuario_username, tipo_mov, categoria, descripcion, moneda, monto = (
                    fila[1], fila[2], fila[3], fila[4], fila[5], fila[6], fila[7], fila[8])
                transacciones.append({
                    'fecha': fecha,
                    'hora': hora,
                    'usuario_username': usuario_username,
                    'tipo_mov': tipo_mov,
                    'categoria': categoria,
                    'descripcion': descripcion,
                    'moneda': moneda,
                    'monto': monto
                })
            except Exception:
                continue
        return transacciones

def insertar_transacciones_importadas(transacciones):
    conn = get_db_connection()
    c = conn.cursor()
    for t in transacciones:
        # Buscar usuario_id por username
        c.execute('SELECT id FROM usuarios WHERE username = ?', (t['usuario_username'],))
        usuario = c.fetchone()
        if usuario:
            usuario_id = usuario[0]
            try:
                c.execute('''
                    INSERT INTO transacciones (fecha, hora, usuario_id, tipo_mov, categoria, descripcion, moneda, monto)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (t['fecha'], t['hora'], usuario_id, t['tipo_mov'], t['categoria'], t['descripcion'], t['moneda'], t['monto']))
            except Exception:
                continue
    conn.commit()
    conn.close()

# ---------------------------------------
# --- INTERFAZ GRÁFICA ------------------
# ---------------------------------------

class Aplicacion(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sistema Financiero")
        self.geometry("1000x600")
        self.resizable(False, False)

        self.usuario_actual = None

        # Inicializar DB
        init_db()

        # Si no hay usuarios, obligar crear primero master
        if not check_first_user_exists():
            self.ventana_registro_usuario(master_creation=True)
        else:
            self.ventana_login()

    # ---------------------------
    # VENTANA LOGIN
    # ---------------------------
    def ventana_login(self):
        self.limpiar_ventana()

        frame = ttk.Frame(self, padding=20)
        frame.pack(expand=True)

        ttk.Label(frame, text="INICIO DE SESIÓN", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)

        ttk.Label(frame, text="Usuario:").grid(row=1, column=0, sticky=tk.E, pady=5)
        self.entry_username = ttk.Entry(frame)
        self.entry_username.grid(row=1, column=1, pady=5)

        ttk.Label(frame, text="Contraseña:").grid(row=2, column=0, sticky=tk.E, pady=5)
        self.entry_password = ttk.Entry(frame, show="*")
        self.entry_password.grid(row=2, column=1, pady=5)

        btn_login = ttk.Button(frame, text="Iniciar sesión", command=self.login)
        btn_login.grid(row=3, column=0, columnspan=2, pady=10)

        btn_registro = ttk.Button(frame, text="Registrarse", command=lambda: self.ventana_registro_usuario(solo_estandar=True))
        btn_registro.grid(row=4, column=0, columnspan=2)

    def login(self):
        username = self.entry_username.get().strip()
        password = self.entry_password.get().strip()

        if not username or not password:
            messagebox.showerror("Error", "Debe ingresar usuario y contraseña")
            return

        usuario = validar_login(username, password)
        if usuario:
            self.usuario_actual = usuario
            self.ventana_principal()
        else:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos")

    # ---------------------------
    # VENTANA REGISTRO USUARIO
    # ---------------------------

    def ventana_registro_usuario(self, master_creation=False, solo_estandar=False):
        self.limpiar_ventana()
        frame = ttk.Frame(self, padding=20)
        frame.pack(expand=True)

        # Título
        if master_creation:
            ttk.Label(frame, text="CREAR USUARIO MASTER (ADMINISTRADOR)", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)
        elif solo_estandar:
            ttk.Label(frame, text="REGISTRAR USUARIO", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)
        else:
            ttk.Label(frame, text="REGISTRAR NUEVO USUARIO", font=("Helvetica", 16)).grid(row=0, column=0, columnspan=2, pady=10)

        # Campos comunes
        ttk.Label(frame, text="Nombre:").grid(row=1, column=0, sticky=tk.E, pady=5)
        entry_nombre = ttk.Entry(frame)
        entry_nombre.grid(row=1, column=1, pady=5)

        ttk.Label(frame, text="Apellido:").grid(row=2, column=0, sticky=tk.E, pady=5)
        entry_apellido = ttk.Entry(frame)
        entry_apellido.grid(row=2, column=1, pady=5)

        ttk.Label(frame, text="Cédula:").grid(row=3, column=0, sticky=tk.E, pady=5)
        entry_cedula = ttk.Entry(frame)
        entry_cedula.grid(row=3, column=1, pady=5)

        ttk.Label(frame, text="Usuario (username):").grid(row=4, column=0, sticky=tk.E, pady=5)
        entry_username = ttk.Entry(frame)
        entry_username.grid(row=4, column=1, pady=5)

        ttk.Label(frame, text="Contraseña:").grid(row=5, column=0, sticky=tk.E, pady=5)
        entry_password = ttk.Entry(frame, show="*")
        entry_password.grid(row=5, column=1, pady=5)

        # Tipo de usuario
        if master_creation:
            tipo_usuario = 'master'
            row_actual = 6
        elif solo_estandar:
            tipo_usuario = 'estandar'
            row_actual = 6
        else:
            ttk.Label(frame, text="Tipo de usuario:").grid(row=6, column=0, sticky=tk.E, pady=5)
            combo_tipo = ttk.Combobox(frame, state="readonly", values=["master", "estandar"])
            combo_tipo.current(1)
            combo_tipo.grid(row=6, column=1, pady=5)
            row_actual = 7

        # Función guardar
        def guardar():
            nombre = entry_nombre.get().strip()
            apellido = entry_apellido.get().strip()
            cedula = entry_cedula.get().strip()
            username = entry_username.get().strip()
            password = entry_password.get().strip()
            tipo = tipo_usuario if master_creation or solo_estandar else combo_tipo.get()

            if not nombre or not apellido or not cedula or not username or not password:
                messagebox.showerror("Error", "Todos los campos son obligatorios")
                return

            exito, error = crear_usuario(nombre, apellido, cedula, username, password, tipo)
            if exito:
                messagebox.showinfo("Éxito", "Usuario creado correctamente")
                if master_creation or solo_estandar:
                    self.ventana_login()
                else:
                    self.ventana_principal()
            else:
                messagebox.showerror("Error", f"No se pudo crear usuario: {error}")

        btn_guardar = ttk.Button(frame, text="Guardar", command=guardar)
        btn_guardar.grid(row=row_actual, column=0, columnspan=2, pady=15)

        # Botón cancelar si no es master_creation
        if not master_creation:
            comando_cancelar = self.ventana_login if solo_estandar else self.ventana_principal
            btn_cancelar = ttk.Button(frame, text="Cancelar", command=comando_cancelar)
            btn_cancelar.grid(row=row_actual + 1, column=0, columnspan=2)


    # ---------------------------
    # VENTANA PRINCIPAL CON PESTAÑAS
    # ---------------------------

    def ventana_principal(self):
        self.limpiar_ventana()
        self.geometry("1000x600")

        menubar = tk.Menu(self)
        self.config(menu=menubar)

        menu_usuario = tk.Menu(menubar, tearoff=0)
        menu_usuario.add_command(label="Cerrar sesión", command=self.cerrar_sesion)
        if self.usuario_actual is not None and self.usuario_actual['tipo'] == 'master':
            menu_usuario.add_command(label="Registrar usuario", command=self.ventana_registro_usuario)
            menu_usuario.add_separator()
            menu_usuario.add_command(label="Importar Excel", command=self.importar_excel)
            menu_usuario.add_command(label="Exportar Excel", command=self.exportar_excel)
            menubar.add_cascade(label=f"Usuario: {self.usuario_actual['username']}", menu=menu_usuario)

        tabControl = ttk.Notebook(self)
        tabControl.pack(expand=1, fill="both")

        # Pestaña para registrar transacciones
        self.tab_registro = ttk.Frame(tabControl)
        tabControl.add(self.tab_registro, text='Registrar Transacción')

        # Pestaña para ver transacciones
        self.tab_ver = ttk.Frame(tabControl)
        tabControl.add(self.tab_ver, text='Ver Transacciones')

        # Pestaña para registrar capital (solo master)
        if self.usuario_actual and self.usuario_actual['tipo'] == 'master':
            self.tab_capital = ttk.Frame(tabControl)
            tabControl.add(self.tab_capital, text='Capital')
            self.construir_tab_capital()

        # Pestaña reportes
        self.tab_reportes = ttk.Frame(tabControl)
        tabControl.add(self.tab_reportes, text='Reportes')

        self.construir_tab_registro()
        self.construir_tab_ver()
        self.construir_tab_reportes()

    def construir_tab_registro(self):
        frame = self.tab_registro
        for widget in frame.winfo_children():
            widget.destroy()

        ttk.Label(frame, text="Tipo de movimiento:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        self.tipo_var = tk.StringVar(value="entrada")
        ttk.Radiobutton(frame, text="Entrada", variable=self.tipo_var, value="entrada").grid(row=0, column=1)
        ttk.Radiobutton(frame, text="Salida", variable=self.tipo_var, value="salida").grid(row=0, column=2)

        ttk.Label(frame, text="Categoría:").grid(row=1, column=0, sticky=tk.W, padx=10)
        self.entry_categoria = ttk.Entry(frame)
        self.entry_categoria.grid(row=1, column=1, columnspan=2, sticky=tk.EW, padx=10)

        ttk.Label(frame, text="Descripción:").grid(row=2, column=0, sticky=tk.W, padx=10)
        self.entry_descripcion = ttk.Entry(frame)
        self.entry_descripcion.grid(row=2, column=1, columnspan=2, sticky=tk.EW, padx=10)

        ttk.Label(frame, text="Moneda:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=10)
        self.moneda_var = tk.StringVar(value="bs")
        ttk.Radiobutton(frame, text="Bolívares (bs)", variable=self.moneda_var, value="bs").grid(row=3, column=1)
        ttk.Radiobutton(frame, text="Dólares (usd)", variable=self.moneda_var, value="usd").grid(row=3, column=2)

        ttk.Label(frame, text="Monto:").grid(row=4, column=0, sticky=tk.W, padx=10)
        self.entry_monto = ttk.Entry(frame)
        self.entry_monto.grid(row=4, column=1, columnspan=2, sticky=tk.EW, padx=10)

        btn_guardar = ttk.Button(frame, text="Registrar", command=self.guardar_transaccion)
        btn_guardar.grid(row=5, column=0, columnspan=3, pady=20)

    def guardar_transaccion(self):
        tipo = self.tipo_var.get()
        categoria = self.entry_categoria.get().strip()
        descripcion = self.entry_descripcion.get().strip()
        moneda = self.moneda_var.get()
        try:
            monto = float(self.entry_monto.get())
        except ValueError:
            messagebox.showerror("Error", "Monto inválido")
            return
        if monto <= 0:
            messagebox.showerror("Error", "El monto debe ser mayor que cero")
            return
        if not categoria:
            messagebox.showerror("Error", "Debe ingresar una categoría")
            return
        if self.usuario_actual:
            registrar_transaccion(self.usuario_actual['id'], tipo, categoria, descripcion, moneda, monto)
            messagebox.showinfo("Éxito", "Transacción registrada")
            self.entry_categoria.delete(0, tk.END)
            self.entry_descripcion.delete(0, tk.END)
            self.entry_monto.delete(0, tk.END)
            self.actualizar_tabla_transacciones()

    def construir_tab_ver(self):
        frame = self.tab_ver
        for widget in frame.winfo_children():
            widget.destroy()

        filtro_frame = ttk.Frame(frame)
        filtro_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(filtro_frame, text="Buscar texto:").grid(row=0, column=0)
        self.busqueda_var = tk.StringVar()
        entry_buscar = ttk.Entry(filtro_frame, textvariable=self.busqueda_var)
        entry_buscar.grid(row=0, column=1, sticky=tk.W)
        entry_buscar.bind('<KeyRelease>', lambda e: self.actualizar_tabla_transacciones())

        ttk.Label(filtro_frame, text="Moneda:").grid(row=0, column=2, padx=5)
        self.filtrar_moneda = ttk.Combobox(filtro_frame, state="readonly", values=["", "bs", "usd"], width=5)
        self.filtrar_moneda.grid(row=0, column=3)
        self.filtrar_moneda.bind('<<ComboboxSelected>>', lambda e: self.actualizar_tabla_transacciones())

        ttk.Label(filtro_frame, text="Tipo:").grid(row=0, column=4, padx=5)
        self.filtrar_tipo = ttk.Combobox(filtro_frame, state="readonly", values=["", "entrada", "salida"], width=8)
        self.filtrar_tipo.grid(row=0, column=5)
        self.filtrar_tipo.bind('<<ComboboxSelected>>', lambda e: self.actualizar_tabla_transacciones())

        columnas = ("id", "fecha", "hora", "usuario", "tipo", "categoria", "descripcion", "moneda", "monto")
        self.tree = ttk.Treeview(frame, columns=columnas, show='headings')
        for col in columnas:
            self.tree.heading(col, text=col.capitalize())
            self.tree.column(col, width=80, anchor=tk.CENTER)
        self.tree.column("descripcion", width=180, anchor=tk.W)
        self.tree.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        self.actualizar_tabla_transacciones()

    def actualizar_tabla_transacciones(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        filtros = {
            'texto': self.busqueda_var.get().strip(),
            'moneda': self.filtrar_moneda.get() if self.filtrar_moneda.get() in ('bs', 'usd') else None,
            'tipo_mov': self.filtrar_tipo.get() if self.filtrar_tipo.get() in ('entrada', 'salida') else None
        }
        transacciones = obtener_transacciones(filtros)
        for t in transacciones:
            self.tree.insert('', 'end', values=t)

    def construir_tab_capital(self):
        frame = self.tab_capital
        for widget in frame.winfo_children():
            widget.destroy()

        ttk.Label(frame, text="Registrar Capital del Inventario").grid(row=0, column=0, columnspan=2, pady=10)

        ttk.Label(frame, text="Monto:").grid(row=1, column=0, sticky=tk.W, padx=10)
        self.entry_monto_cap = ttk.Entry(frame)
        self.entry_monto_cap.grid(row=1, column=1, columnspan=2, sticky=tk.EW, padx=10)

        ttk.Label(frame, text="Descripción:").grid(row=2, column=0, sticky=tk.W, padx=10)
        self.entry_descripcion_cap = ttk.Entry(frame)
        self.entry_descripcion_cap.grid(row=2, column=1, columnspan=2, sticky=tk.EW, padx=10)

        btn_guardar = ttk.Button(frame, text="Registrar", command=self.guardar_capital)
        btn_guardar.grid(row=3, column=0, columnspan=2, pady=10)
    
    def guardar_capital(self):
        try:
            monto = float(self.entry_monto_cap.get())
        except ValueError:
            messagebox.showerror("Error", "Monto inválido")
            return
        descripcion = self.entry_descripcion_cap.get().strip()
        if monto <= 0:
            messagebox.showerror("Error", "El monto debe ser mayor a cero")
            return
        if self.usuario_actual:
            registrar_transaccion(self.usuario_actual['id'], "entrada", "inventario", descripcion, "bs", monto)
            messagebox.showinfo("Éxito", "Capital del inventario registrado")
            self.entry_monto_cap.delete(0, tk.END)
            self.entry_descripcion_cap.delete(0, tk.END)
            self.actualizar_tabla_transacciones()
            self.actualizar_balance()

    def construir_tab_reportes(self):
        frame = self.tab_reportes
        for widget in frame.winfo_children():
            widget.destroy()

        ttk.Label(frame, text="Generar Reporte PDF de transacciones", font=("Helvetica", 14)).pack(pady=10)

        btn_generar = ttk.Button(frame, text="Generar Reporte PDF", command=self.generar_reporte_pdf_ui)
        btn_generar.pack(pady=5)

        ttk.Label(frame, text="Exportar todas las transacciones a Excel", font=("Helvetica", 14)).pack(pady=10)

        btn_exportar = ttk.Button(frame, text="Exportar a Excel", command=self.exportar_excel)
        btn_exportar.pack(pady=5)

        if self.usuario_actual and self.usuario_actual['tipo'] == 'master':
            ttk.Label(frame, text="Importar transacciones desde Excel", font=("Helvetica", 14)).pack(pady=10)
            btn_importar = ttk.Button(frame, text="Importar desde Excel", command=self.importar_excel)
            btn_importar.pack(pady=5)

        ttk.Label(frame, text="Balance Actual:", font=("Helvetica", 14)).pack(pady=10)
        self.label_balance = ttk.Label(frame, text="")
        self.label_balance.pack(pady=5)

        self.actualizar_balance()

    def actualizar_balance(self):
        balances = calcular_balance()
        texto = f"Bolívares (bs): {balances['bs']:.2f}  |  Dólares (usd): {balances['usd']:.2f}"
        self.label_balance.config(text=texto)

    def generar_reporte_pdf_ui(self):
        transacciones = obtener_transacciones()
        if not transacciones:
            messagebox.showinfo("Info", "No hay transacciones para generar reporte")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivo PDF", "*.pdf")])
        if not ruta:
            return
        generar_reporte_pdf(ruta, transacciones, self.usuario_actual)
        messagebox.showinfo("Éxito", f"Reporte PDF generado:\n{ruta}")

    def exportar_excel(self):
        if self.usuario_actual and self.usuario_actual['tipo'] != 'master':
            messagebox.showerror("Permiso denegado", "Solo usuarios master pueden exportar")
            return
        transacciones = obtener_transacciones()
        if not transacciones:
            messagebox.showinfo("Info", "No hay transacciones para exportar")
            return
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        exportar_a_excel(ruta, transacciones)
        messagebox.showinfo("Éxito", f"Archivo Excel exportado:\n{ruta}")

    def importar_excel(self):
        if self.usuario_actual and self.usuario_actual['tipo'] != 'master':
            messagebox.showerror("Permiso denegado", "Solo usuarios master pueden importar")
            return
        ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        transacciones = importar_desde_excel(ruta)
        if not transacciones:
            messagebox.showerror("Error", "Archivo Excel inválido o sin datos")
            return
        insertar_transacciones_importadas(transacciones)
        messagebox.showinfo("Éxito", "Transacciones importadas correctamente")
        self.actualizar_tabla_transacciones()
        self.actualizar_balance()

    def cerrar_sesion(self):
        self.usuario_actual = None
        self.ventana_login()

    def limpiar_ventana(self):
        for widget in self.winfo_children():
            widget.destroy()


if __name__ == '__main__':
    app = Aplicacion()
    app.mainloop()
